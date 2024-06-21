import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side
import os
import math
import numpy as np

current_directory = os.path.dirname(__file__)
path_plantilla = os.path.join(current_directory, '..', 'data' ,'catastro.xlsx')
path_project_cantv = os.path.join(current_directory, '..', 'data' , 'edificaciones_bellomonte.xlsx')
path_capa_puntos = os.path.join(current_directory, '..', 'data', 'capa_puntos_bellomonte.xlsx')

name_sheet_edificaciones = "EDIFICACIONES"
list_column_origin = ['cod_parc', 'parroquia', 'Tipo de Vía', 'Dirección', 'Uso Inmueble', 'Nombre Inmueble', 'Tipo Inmueble', 'Fecha', 'Total Unidades', 'Total Residenciales', 
 'Total Comerciales', 'Total Oficinas', 'Total Medico Asistencial', 'Total Gubernamental', 'Total Educativo', 'Manzana asignada', 'Total Religioso', 'Total construccion', 
 'Total abandonadas', 'Total Plaza', 'Total Parques', 'sectores']

list_column_origin_postes = ['Punto de referencia', 'Tipo de Poste', 'El poste tiene código', 'Fecha Levantaamiento', 'Nro. Poste', 'Manzana Asignada', 'sector']
sheet_postes_origin = 'POSTES'


sheet_tanquilla_origin = "TANQUILLAS"
list_column_origin_tanquilla = ['Punto de Referencia', 'Tipo de tanquilla',	'Fecha Levantamiento', 'sector',	'Manzana asignada']

#Relaciones de archivos y columnas.
#Proyecto_CANTV(EDIFICACIONES) -> catastro(Demanda)

values_planta_externa = {
  "Manzana asignada":"Manzana",
  "cod_parc": "cod_parc",
  "Tipo Inmueble": "Tipo Inmueble",
  "Uso Inmueble": "Uso Inmueble",
  "Tipo de Vía": "Tipo de Vía",
  "Dirección": "Dirección",
  "sectores": "sectores",
  "Nombre Inmueble": "Nombre Inmueble"
} 

values_demanda = {
  "sectores":"Area",
  "Manzana asignada":"Manzana", 
  "Nro ID Plano": "Nro ID Plano", 
  "Tipo de Inmueble": {
    "Edificio":"Edf",
    "Casa":"Casa",
    "Quinta":"Qta",
    "Quintas Pareadas":"Qta Pareada",	
    "Conjunto de Quintas":"Conjunto Qta",	
    "Conjunto de Edificios":"Conjunto Edif",	
    "Local Comercial":"Local", 
    "Centro Comercial":"centro comercial",
    "Oficina": "oficina",  	
    "Club":"club",	
    "Colegio":"colegio",	
    "Galpon":"galpon",
    "Templo":"templo",
    "Universidad":"universidad",
    "Parque":"parque",
    "Plaza":"plaza",
    "Cancha":"cancha",
    "Estacion de Servicio":"estacion de servicio",
    "Zona Militar":"zona militar",
    "Campo Santo":"campo santo",
    "Caseta":"caseta",
    "Estación de Metro":"estacion de metro",
    "Estacion de Metro":"estacion de metro",
    "Plaza de Toros":"plaza de toros",
    "Puertos":"puertos",
    "Teatro": "teatro",
    "Terreno": "terreno",
    "Abandonado":"abandonado",
    "Centro Empresarial":"centro empresarial",
    "Aeropuerto":"aeropuerto",
    "Construcción":"construccion",
    "Construccion":"construccion"
   
  },
  "Total Residenciales": "Nº Residenciales",
  "Total Comerciales": "Nº Comerciales",
  "Total Oficinas": "Nº Oficinas",
  "Total Medico Asistencial": "Nº Medico Asistencial",
  "Total Gubernamental": "Nº Gubernamental",
  "Total Educativo": "Nº Educativo",
  "Total Religioso": "Nº Religioso",
  "Total construccion": "Nº construccion",
  "Total abandonadas": "Nº abandonadas",
  "Total Plaza": "Nº Plaza",
  "Total Parques": "Nº Parques",
  "Dirección":"DIRECCIÒN",
  "Nombre Inmueble":"NOMBRE DEL INMUEBLE",
}

tipes_tanquilla = {
  'Tipo A (Rectangular)': 'Tanquilla Tipo A',
  'Tipo B (Cuadrada)': 'Tanquilla Tipo B',
  'Tanque (Redonda)': 'Tanque (Redonda)',
  'Tanquillon': 'Tanquillón',
  'Armario': 'ARMARIO',
  'Terminal': 'TERMINAL'
}

tipes_postes = {
  'Poste eléctrico de concreto': 'PEC',
  'Poste eléctrico de hierro': 'PEH',
  'Poste de cemento con transformador': 'PECT',
  'Poste Eléctrico de hierro con transformador': 'PEHT',
  'Poste de CANTV': 'PT',
  'Poste de telecomunicaciones': 'PTELCO',
  'Poste de Telefónica': 'PTELEF',
  'Poste de Digitel': 'PD',
  'Poste del 911': 'P911',
  'Poste de alumbrado público': 'PAP'
}

def get_type_property(type_property, conpare_property):
  object_type = values_demanda["Tipo de Inmueble"]
  if(type_property in object_type):
    value_type_property = object_type[type_property]
    if(conpare_property.lower() == value_type_property.lower()):
      return "X"
    else:
      return ""
  else:
    return ""

def get_type_property_postes(type_property, conpare_property):
  if(type_property in tipes_postes):
    value_type_property = tipes_postes[type_property]
    if(conpare_property.lower() == value_type_property.lower()):
      return "X"
    else:
      return ""
  else:
    return ""

def get_type_property_sub(type_property, conpare_property):
  if(type_property in tipes_tanquilla):
    value_type_property = tipes_tanquilla[type_property]
    if(conpare_property.lower() == value_type_property.lower()):
      return "X"
    else:
      return ""
  else:
    return ""

def get_number(num):
  if(math.isnan(num)):
    return ""
  else:
    if(num == 0):
      return ""
    return str(num)

def get_all_sectors(path):
  excel_file = pd.ExcelFile(path)
  index = 0
  sheet_name = excel_file.sheet_names[index]
  df = read_file(path,  ['sectores'], sheet_name=sheet_name)
  
  return df['sectores'].toList()

def read_file(path_file, list_column, name_sheet="Hoja1"):
  try:

    if(len(list_column) > 0):
      df = pd.read_excel(path_file, sheet_name=name_sheet, usecols=list_column, header=0)
    else:
      df = pd.read_excel(path_file, sheet_name=name_sheet)
    return df
  except Exception as e:
    print(f"Algo salio mal: {e}")


def get_all_row_area(df):
  try:
    list_rows_file = []
    list_rows_planta_externa = []
    
    for index, row in df.iterrows():
      new_row = [
        row['Manzana asignada'],
        "",
        get_type_property(row["Tipo Inmueble"], "Edf"),
        get_type_property(row["Tipo Inmueble"], "Casa"),
        get_type_property(row["Tipo Inmueble"], "Qta"),
        get_type_property(row["Tipo Inmueble"], "Qta Pareada"),
        get_type_property(row["Tipo Inmueble"], "Conjunto Qta"),
        get_type_property(row["Tipo Inmueble"],"Conjunto Edif"),
        get_type_property(row["Tipo Inmueble"],"Local"),
        get_type_property(row["Tipo Inmueble"],"centro comercial"),
        get_type_property(row["Tipo Inmueble"],"oficina"),
        get_type_property(row["Tipo Inmueble"],"club"),
        get_type_property(row["Tipo Inmueble"],"colegio"),
        get_type_property(row["Tipo Inmueble"],"galpon"),
        get_type_property(row["Tipo Inmueble"],"templo"),
        get_type_property(row["Tipo Inmueble"],"universidad"),
        get_type_property(row["Tipo Inmueble"],"parque"),
        get_type_property(row["Tipo Inmueble"],"plaza"),
        get_type_property(row["Tipo Inmueble"],"cancha"),
        get_type_property(row["Tipo Inmueble"],"estacion de servicio"),
        get_type_property(row["Tipo Inmueble"],"zona militar"),        
        get_type_property(row["Tipo Inmueble"],"campo santo"),
        get_type_property(row["Tipo Inmueble"],"caseta"),
        get_type_property(row["Tipo Inmueble"],"estacion de metro"),
        get_type_property(row["Tipo Inmueble"],"plaza de toros"),
        get_type_property(row["Tipo Inmueble"],"puertos"),
        get_type_property(row["Tipo Inmueble"],"teatro"),
        get_type_property(row["Tipo Inmueble"],"terreno"),
        get_type_property(row["Tipo Inmueble"],"abandonado"),
        get_type_property(row["Tipo Inmueble"],"centro empresarial"),
        get_type_property(row["Tipo Inmueble"],"aeropuerto"),
        get_type_property(row["Tipo Inmueble"],"construccion"),
        get_number(row['Total Residenciales']),
        get_number(row['Total Comerciales']),
        get_number(row['Total Oficinas']),
        get_number(row['Total Medico Asistencial']),
        get_number(row['Total Gubernamental']),
        get_number(row['Total Educativo']),
        get_number(row['Total Religioso']),
        get_number(row['Total construccion']),
        get_number(row['Total abandonadas']),
        get_number(row['Total Plaza']),
        get_number(row['Total Parques']),
        ", ".join([str(row[column]) for column in ['parroquia', 'sectores', 'Dirección'] if pd.notna(row[column])]),
        row['Nombre Inmueble']
      ]
      list_rows_file.append(new_row)
    #str(row['parroquia']) + ", " +  str(row['sectores']) + ", " + str(row['Dirección']),
    for index, row in df.iterrows():
      new_row = [
        row['Manzana asignada'],
        row['cod_parc'],
        row['Tipo Inmueble'],
        row['Uso Inmueble'],
        row['Tipo de Vía'],
        ", ".join([str(row[column]) for column in ['parroquia', 'sectores', 'Dirección'] if pd.notna(row[column])]),
        row['Nombre Inmueble']
      ]
      list_rows_planta_externa.append(new_row)
    
    #str(row['parroquia']) + ", " +  str(row['sectores']) + ", " + str(row['Dirección']),

    list_rows_file = sorted(list_rows_file, key=lambda x: x[0])
    list_rows_planta_externa = sorted(list_rows_planta_externa, key=lambda x : x[0])
    return list_rows_file, list_rows_planta_externa
  except Exception as e:
    print(f"Algo salio mal: {e}")

def get_code(codigo, conpare):
  if isinstance(codigo, str):
    if(codigo.lower() == conpare):
      return 'X'
    return ''    
  return ''

def get_all_row_pdr(df_postes, df_sub):
  try:
    list_row_postes = []
    list_row_tanquilla = []

    for index, row in df_postes.iterrows():
      new_row = [
        row['Manzana Asignada'],
        '',
        get_type_property_postes(row['Tipo de Poste'], 'PEC'),
        get_type_property_postes(row['Tipo de Poste'], 'PEH'),
        get_type_property_postes(row['Tipo de Poste'], 'PECT'),
        get_type_property_postes(row['Tipo de Poste'], 'PEHT'),
        get_type_property_postes(row['Tipo de Poste'], 'PT'),
        get_type_property_postes(row['Tipo de Poste'], 'PTELCO'),
        get_type_property_postes(row['Tipo de Poste'], 'PTELEF'),
        get_type_property_postes(row['Tipo de Poste'], 'PD'),
        get_type_property_postes(row['Tipo de Poste'], 'P911'),
        get_type_property_postes(row['Tipo de Poste'], 'PAP'),
        get_code(row['El poste tiene código'], 'si'),
        get_code(row['El poste tiene código'], 'no'),
        '',
        str(row['sector']) + ", " + str(row['Punto de referencia']),
        #", ".join([str(row[column]) for column in ['sector', 'Punto de referencia'] if pd.notna(row[column])]),
        'Poste'
      ]
      list_row_postes.append(new_row)
      #str(row['sector']) + ", " + str(row['Punto de referencia']),
    for index, row in df_sub.iterrows():
      new_row = [
        row['Manzana asignada'],
        '',        
        get_type_property_sub(row['Tipo de tanquilla'], 'Tanquilla Tipo A'),
        get_type_property_sub(row['Tipo de tanquilla'], 'Tanquilla Tipo B'),
        get_type_property_sub(row['Tipo de tanquilla'], 'Tanque (Redonda)'),
        get_type_property_sub(row['Tipo de tanquilla'], 'Tanquillón'),
        get_type_property_sub(row['Tipo de tanquilla'], 'ARMARIO'),
        get_type_property_sub(row['Tipo de tanquilla'], 'TERMINAL'),
        #", ".join([str(row[column]) for column in ['sector', 'Punto de referencia'] if pd.notna(row[column])]),
        str(row['sector']) + ", " + str(row['Punto de Referencia']),
        'Subterraneo'
      ]
      list_row_tanquilla.append(new_row)
      
    return list_row_postes, list_row_tanquilla
  except Exception as e:
    print(f'get_all_row_pdr Algo salio mal {e}')

def create_one_exel_area(area, path_file, list_column , name_sheet, sheet_postes_origin, sheet_tanquilla_origin, capa_puntos_path):
  try:
    #df_Demanda
    df = read_file(path_file, list_column , name_sheet)
    df = df[df['sectores'].str.lower() == area.lower() ]
    
    #df_Postes
    df_postes = read_file(capa_puntos_path, list_column_origin_postes, sheet_postes_origin)
    df_sub = read_file(capa_puntos_path, list_column_origin_tanquilla, sheet_tanquilla_origin)

    df_postes = df_postes[df_postes['sector'].str.lower() == area.lower()]
    df_sub = df_sub[df_sub['sector'].str.lower() == area.lower()]

    list_row_postes, list_row_tanquilla = get_all_row_pdr(df_postes, df_sub)
    all_row_new_file, all_row_planta_externa = get_all_row_area(df)
    
    plantilla = openpyxl.load_workbook(path_plantilla)
    shell_plantilla = plantilla['Demanda']
    shell_plantilla_postes = plantilla['PDR POTES']
    shell_plantilla_tanquilla = plantilla['PDR TANQUILLAS']
    shell_plantilla_PLANTA_EXTERNA = plantilla['PLANTA EXTERNA']

    shell_plantilla['B7'] = area
    shell_plantilla_postes['B4'] = area
    shell_plantilla_tanquilla['B4'] = area

    init_row_planta_externa = 5
    init_row_pdr = 7
    init_row = 10

    border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

    for i, row in enumerate(all_row_new_file, start = init_row):
      for j, value in enumerate(row, start=1):
        if(j == 1):
          if isinstance(value, (float)):
            value = str(int(value))
          elif np.isnan(value):
            value = 'null'
          if(value == 0):
            value = "0"
          else:
             value = str(value)

          if(len(value) <= 1):
            value = "00" + value
          elif(len(value) <= 2 ):
            value = "0" + value

        cell = shell_plantilla.cell(row = i, column = j, value= value)
        cell.border = border
    
    for i, row in enumerate(list_row_postes, start = init_row_pdr):
      for j, value in enumerate(row, start=1):
        if(j == 1):
          if(value == 0):
            value = "0"
          else:
             value = str(value)

          value = str(value)
          if(len(value) <= 1):
            value = "00" + value
          elif(len(value) <= 2 ):
            value = "0" + value

        cell = shell_plantilla_postes.cell(row = i, column = j, value= value)
        cell.border = border

    for i, row in enumerate(list_row_tanquilla, start = init_row_pdr):
      for j, value in enumerate(row, start=1):
        if(j == 1):
          if(value == 0):
            value = "0"
          else:
             value = str(value)

          value = str(value)
          if(len(value) <= 1):
            value = "00" + value
          elif(len(value) <= 2 ):
            value = "0" + value

        cell = shell_plantilla_tanquilla.cell(row = i, column = j, value= value)
        cell.border = border

    for i, row in enumerate(all_row_planta_externa, start = init_row_planta_externa):
      for j, value in enumerate(row, start=1):
        if(j == 1):
          if(value == 0):
            value = "0"
          else:
             value = str(value)

          value = str(value)
          if(len(value) <= 1):
            value = "00" + value
          elif(len(value) <= 2 ):
            value = "0" + value

        cell = shell_plantilla_PLANTA_EXTERNA.cell(row = i, column = j, value= value)
        cell.border = border

    path_save = os.path.join(current_directory, '..', 'data' ,'new_catastro.xlsx')
    plantilla.save(path_save)
    return
  except Exception as e:
    print(f"Algo salio mal: {e}")

""" "parsing-data-cantv/parsing_data_cantv/data/catastro.xlsx" """
def proccessData():
  create_one_exel_area('Santa Monica', path_project_cantv, list_column_origin , 'EDIFICACIONES')

#create_one_exel_area('COLINAS DE BELLO MONTE', path_project_cantv, list_column_origin , 'EDIFICACIONES', sheet_postes_origin, sheet_tanquilla_origin, path_capa_puntos)
#def set_value_cells(value_origin)